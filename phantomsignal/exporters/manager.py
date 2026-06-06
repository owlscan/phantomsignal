"""
PhantomSignal Export Manager — Intelligence Packet Compiler
Handles all export formats with optional compression and encryption.

Author:  the-clipper
AI:      Claude (Anthropic)
License: MIT — see LICENSE
"""
from __future__ import annotations

import gzip
import hashlib
import json
import logging
import os
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from phantomsignal.core.config import config as phantomsignal_config
from phantomsignal.core.database import get_db
from phantomsignal.core.models import Export, Scan, ScanResult

logger = logging.getLogger("phantomsignal.exporters")


class ExportManager:
    """Central export coordinator — compiles, compresses, and optionally encrypts intel packets."""

    SUPPORTED_FORMATS = {
        "json": ".json",
        "csv": ".csv",
        "pdf": ".pdf",
        "html": ".html",
        "xml": ".xml",
        "xlsx": ".xlsx",
        "stix": ".json",
        "markdown": ".md",
    }

    def __init__(self, output_dir: Optional[str] = None):
        self.output_dir = Path(output_dir or phantomsignal_config.get("export", "output_dir", default="/tmp")).resolve()
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export(
        self,
        scan_id: str,
        fmt: str = "json",
        compress: bool = False,
        encrypt: bool = False,
        encryption_password: Optional[str] = None,
        include_raw: bool = False,
    ) -> Dict[str, Any]:
        """
        Export a scan's intelligence packet to the requested format.
        Returns metadata about the created export file.
        """
        with get_db() as db:
            scan = db.query(Scan).filter(Scan.id == scan_id).first()
            if not scan:
                raise ValueError(f"Scan {scan_id} not found in the grid.")
            results = db.query(ScanResult).filter(ScanResult.scan_id == scan_id).all()

            scan_dict = scan.to_dict()
            results_list = [r.to_dict() for r in results]

        if fmt not in self.SUPPORTED_FORMATS:
            raise ValueError(f"Unsupported format: {fmt}. Available: {list(self.SUPPORTED_FORMATS.keys())}")

        ext = self.SUPPORTED_FORMATS[fmt]
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        safe_target = "".join(c if c.isalnum() or c in ".-_" else "_" for c in scan_dict["target"])[:40]
        base_name = f"phantomsignal_{safe_target}_{timestamp}"
        output_path = self.output_dir / f"{base_name}{ext}"

        payload = {
            "phantomsignal_version": "1.0.0",
            "export_timestamp": datetime.utcnow().isoformat(),
            "scan": scan_dict,
            "results": results_list,
            "result_count": len(results_list),
            "export_format": fmt,
        }

        # Generate the export file
        exporter = self._get_exporter(fmt)
        exporter(payload, output_path)

        file_size = output_path.stat().st_size
        final_path = output_path

        # Compression
        if compress:
            final_path = self._compress(output_path)
            if final_path != output_path:
                output_path.unlink(missing_ok=True)
            file_size = final_path.stat().st_size

        # Encryption
        if encrypt and encryption_password:
            final_path = self._encrypt_file(final_path, encryption_password)
            file_size = final_path.stat().st_size

        checksum = self._compute_checksum(final_path)

        with get_db() as db:
            export_record = Export(
                scan_id=scan_id,
                format=fmt,
                file_path=str(final_path),
                file_size=file_size,
                compressed=compress,
                encrypted=encrypt,
                checksum=checksum,
            )
            db.add(export_record)

        return {
            "file_path": str(final_path),
            "file_name": final_path.name,
            "format": fmt,
            "file_size": file_size,
            "file_size_human": self._human_size(file_size),
            "compressed": compress,
            "encrypted": encrypt,
            "checksum_sha256": checksum,
            "result_count": len(results_list),
            "timestamp": datetime.utcnow().isoformat(),
        }

    def _get_exporter(self, fmt: str):
        exporters = {
            "json": self._export_json,
            "csv": self._export_csv,
            "pdf": self._export_pdf,
            "html": self._export_html,
            "xml": self._export_xml,
            "xlsx": self._export_xlsx,
            "stix": self._export_stix,
            "markdown": self._export_markdown,
        }
        return exporters[fmt]

    def _export_json(self, payload: Dict, path: Path) -> None:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, default=str, ensure_ascii=False)

    def _export_csv(self, payload: Dict, path: Path) -> None:
        import csv
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["scan_id", "target", "module", "type", "source",
                             "confidence", "relevance_score", "timestamp", "data_summary", "tags"])
            for r in payload["results"]:
                data = r.get("data", {})
                summary = json.dumps(data, default=str)[:500]
                writer.writerow([
                    r.get("scan_id"), payload["scan"].get("target"),
                    r.get("module"), r.get("result_type"), r.get("source"),
                    r.get("confidence"), r.get("relevance_score"),
                    r.get("timestamp"), summary,
                    ",".join(r.get("tags", [])),
                ])

    def _export_xml(self, payload: Dict, path: Path) -> None:
        from xml.etree.ElementTree import Element, SubElement, ElementTree, indent
        root = Element("PhantomSignalReport")
        root.set("version", payload["phantomsignal_version"])
        root.set("exported", payload["export_timestamp"])

        scan_el = SubElement(root, "Scan")
        for k, v in payload["scan"].items():
            el = SubElement(scan_el, k)
            el.text = str(v) if v is not None else ""

        results_el = SubElement(root, "Results")
        results_el.set("count", str(payload["result_count"]))
        for r in payload["results"]:
            result_el = SubElement(results_el, "Result")
            for k in ["module", "result_type", "source", "confidence", "timestamp"]:
                if k in r:
                    el = SubElement(result_el, k)
                    el.text = str(r[k])
            data_el = SubElement(result_el, "Data")
            data_el.text = json.dumps(r.get("data", {}), default=str)

        tree = ElementTree(root)
        indent(tree)
        tree.write(path, encoding="unicode", xml_declaration=True)

    def _export_xlsx(self, payload: Dict, path: Path) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment
        except ImportError:
            self._export_csv(payload, path.with_suffix(".csv"))
            return

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"

        header_fill = PatternFill("solid", fgColor="0A0A0F")
        header_font = Font(color="00FF41", bold=True)

        scan = payload["scan"]
        ws_summary.append(["PhantomSignal Intelligence Report"])
        ws_summary.append(["Target", scan.get("target")])
        ws_summary.append(["Scan Type", scan.get("scan_type")])
        ws_summary.append(["Status", scan.get("status")])
        ws_summary.append(["Shadow Score", scan.get("shadow_score")])
        ws_summary.append(["Threat Level", scan.get("threat_level")])
        ws_summary.append(["Results", payload["result_count"]])
        ws_summary.append(["Exported", payload["export_timestamp"]])

        ws_results = wb.create_sheet("Results")
        headers = ["Module", "Type", "Source", "Confidence", "Relevance", "Timestamp", "Anomaly", "Tags", "Data"]
        ws_results.append(headers)
        for cell in ws_results[1]:
            cell.fill = header_fill
            cell.font = header_font

        for r in payload["results"]:
            ws_results.append([
                r.get("module"), r.get("result_type"), r.get("source"),
                r.get("confidence"), r.get("relevance_score"), r.get("timestamp"),
                r.get("is_anomaly"), ",".join(r.get("tags", [])),
                json.dumps(r.get("data", {}), default=str)[:500],
            ])

        wb.save(path)

    def _export_pdf(self, payload: Dict, path: Path) -> None:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        except ImportError:
            self._export_json(payload, path.with_suffix(".json"))
            return

        doc = SimpleDocTemplate(str(path), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle(
            "PhantomSignalTitle",
            parent=styles["Title"],
            textColor=colors.HexColor("#00FF41"),
            backColor=colors.HexColor("#0A0A0F"),
            fontSize=24,
        )
        story.append(Paragraph("PhantomSignal OSINT Report", title_style))
        story.append(Spacer(1, 0.5 * cm))

        scan = payload["scan"]
        meta_data = [
            ["Field", "Value"],
            ["Target", scan.get("target", "")],
            ["Scan Type", scan.get("scan_type", "")],
            ["Status", scan.get("status", "")],
            ["Shadow Score", str(scan.get("shadow_score", ""))],
            ["Threat Level", scan.get("threat_level", "")],
            ["Duration", f"{scan.get('duration_seconds', 0):.1f}s"],
            ["Results", str(payload["result_count"])],
            ["Exported", payload["export_timestamp"]],
        ]

        table = Table(meta_data, colWidths=[5 * cm, 12 * cm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0A0F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#00FF41")),
            ("FONTNAME", (0, 0), (-1, 0), "Courier-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#1A1A2E")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#0D0D1A"), colors.HexColor("#0A0A0F")]),
            ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#C0C0D0")),
        ]))
        story.append(table)
        story.append(Spacer(1, cm))

        story.append(Paragraph("Intelligence Results", styles["Heading2"]))
        for r in payload["results"][:200]:
            data_str = json.dumps(r.get("data", {}), default=str, indent=2)[:800]
            result_text = f"<b>[{r.get('module', '')}]</b> {r.get('result_type', '')} | {r.get('source', '')} | conf: {r.get('confidence', '')}"
            story.append(Paragraph(result_text, styles["Normal"]))
            story.append(Spacer(1, 0.2 * cm))

        doc.build(story)

    def _export_html(self, payload: Dict, path: Path) -> None:
        scan = payload["scan"]
        results_html = ""
        for r in payload["results"]:
            anomaly_class = "anomaly" if r.get("is_anomaly") else ""
            data_str = json.dumps(r.get("data", {}), indent=2, default=str)
            results_html += f"""
            <div class="result {anomaly_class}">
                <div class="result-header">
                    <span class="module">[{r.get('module', '')}]</span>
                    <span class="type">{r.get('result_type', '')}</span>
                    <span class="source">⚡ {r.get('source', '')}</span>
                    <span class="confidence">CONF: {r.get('confidence', 0):.0%}</span>
                    {('<span class="anomaly-badge">⚠ ANOMALY</span>' if r.get("is_anomaly") else "")}
                </div>
                <pre class="data">{data_str}</pre>
            </div>"""

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>PhantomSignal Report — {scan.get('target')}</title>
<style>
:root {{ --neon-green: #00ff41; --neon-cyan: #00f3ff; --neon-purple: #b026ff; --dark: #0a0a0f; --card: #0d0d1a; }}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ background: var(--dark); color: #c0c0d0; font-family: 'Courier New', monospace; padding: 2rem; }}
h1 {{ color: var(--neon-green); font-size: 2rem; text-shadow: 0 0 20px var(--neon-green); margin-bottom: 1rem; }}
h2 {{ color: var(--neon-cyan); margin: 1.5rem 0 0.5rem; }}
.meta {{ background: var(--card); border: 1px solid #1a1a2e; padding: 1rem; border-radius: 4px; margin-bottom: 1.5rem; }}
.meta table {{ width: 100%; border-collapse: collapse; }}
.meta td {{ padding: 0.3rem 0.5rem; border-bottom: 1px solid #1a1a2e; }}
.meta td:first-child {{ color: var(--neon-green); width: 40%; }}
.result {{ background: var(--card); border: 1px solid #1a1a2e; padding: 1rem; margin: 0.5rem 0; border-radius: 4px; }}
.result.anomaly {{ border-color: #ff2d55; box-shadow: 0 0 10px rgba(255,45,85,0.3); }}
.result-header {{ display: flex; gap: 1rem; flex-wrap: wrap; margin-bottom: 0.5rem; align-items: center; }}
.module {{ color: var(--neon-green); font-weight: bold; }}
.type {{ color: var(--neon-cyan); }}
.source {{ color: #9b59b6; }}
.confidence {{ color: #ff6b00; font-size: 0.8rem; }}
.anomaly-badge {{ background: #ff2d55; color: white; padding: 0.1rem 0.4rem; border-radius: 3px; font-size: 0.8rem; }}
pre.data {{ font-size: 0.75rem; overflow-x: auto; padding: 0.5rem; background: #05050a; border-radius: 3px; color: #a0a0b0; max-height: 300px; overflow-y: auto; }}
.shadow-score {{ font-size: 3rem; color: var(--neon-green); text-shadow: 0 0 30px var(--neon-green); }}
footer {{ margin-top: 3rem; border-top: 1px solid #1a1a2e; padding-top: 1rem; color: #404050; text-align: center; }}
</style>
</head>
<body>
<h1>PhantomSignal Intelligence Report</h1>
<div class="meta">
  <table>
    <tr><td>Target</td><td>{scan.get('target')}</td></tr>
    <tr><td>Scan Type</td><td>{scan.get('scan_type')}</td></tr>
    <tr><td>Shadow Score</td><td><span class="shadow-score">{scan.get('shadow_score', 0):.0f}</span>/100</td></tr>
    <tr><td>Threat Level</td><td>{scan.get('threat_level', 'UNKNOWN').upper()}</td></tr>
    <tr><td>Results</td><td>{payload['result_count']}</td></tr>
    <tr><td>Duration</td><td>{scan.get('duration_seconds', 0):.1f}s</td></tr>
    <tr><td>Generated</td><td>{payload['export_timestamp']}</td></tr>
  </table>
</div>
<h2>Intelligence Findings ({payload['result_count']} signals)</h2>
{results_html}
<footer>Generated by PhantomSignal OSINT Framework v{payload['phantomsignal_version']} | "See everything. Leave no trace."</footer>
</body>
</html>"""
        with open(path, "w", encoding="utf-8") as f:
            f.write(html)

    def _export_stix(self, payload: Dict, path: Path) -> None:
        """Export as STIX 2.1 bundle for threat intelligence platforms."""
        import uuid
        scan = payload["scan"]
        objects = []

        # Create STIX identity for report source
        identity = {
            "type": "identity",
            "spec_version": "2.1",
            "id": f"identity--{uuid.uuid4()}",
            "name": "PhantomSignal OSINT Framework",
            "identity_class": "tool",
        }
        objects.append(identity)

        # Create indicators from results
        for r in payload["results"]:
            data = r.get("data", {})
            rtype = r.get("result_type", "")

            if "ip" in rtype and data.get("ip"):
                objects.append({
                    "type": "indicator",
                    "spec_version": "2.1",
                    "id": f"indicator--{uuid.uuid4()}",
                    "pattern": f"[ipv4-addr:value = '{data['ip']}']",
                    "pattern_type": "stix",
                    "valid_from": r.get("timestamp", datetime.utcnow().isoformat()),
                    "name": f"IP: {data['ip']}",
                    "indicator_types": ["malicious-activity" if r.get("is_anomaly") else "unknown"],
                    "confidence": int((r.get("confidence", 0.5)) * 100),
                })
            elif "domain" in rtype and data.get("domain"):
                objects.append({
                    "type": "domain-name",
                    "spec_version": "2.1",
                    "id": f"domain-name--{uuid.uuid4()}",
                    "value": data["domain"],
                })

        bundle = {
            "type": "bundle",
            "id": f"bundle--{uuid.uuid4()}",
            "spec_version": "2.1",
            "objects": objects,
        }

        with open(path, "w", encoding="utf-8") as f:
            json.dump(bundle, f, indent=2, default=str)

    def _export_markdown(self, payload: Dict, path: Path) -> None:
        scan = payload["scan"]
        md = f"""# PhantomSignal OSINT Report

> **Target:** `{scan.get('target')}`
> **Shadow Score:** {scan.get('shadow_score', 0):.0f}/100
> **Threat Level:** {scan.get('threat_level', 'UNKNOWN').upper()}
> **Results:** {payload['result_count']}
> **Generated:** {payload['export_timestamp']}

---

## Scan Metadata

| Field | Value |
|-------|-------|
| Scan ID | `{scan.get('id')}` |
| Type | {scan.get('scan_type')} |
| Status | {scan.get('status')} |
| Duration | {scan.get('duration_seconds', 0):.1f}s |
| Profile | {scan.get('profile')} |

---

## Intelligence Findings

"""
        for r in payload["results"]:
            anomaly_marker = " ⚠️ **ANOMALY**" if r.get("is_anomaly") else ""
            md += f"### [{r.get('module', '')}] {r.get('result_type', '')}{anomaly_marker}\n"
            md += f"- **Source:** {r.get('source', '')}  \n"
            md += f"- **Confidence:** {r.get('confidence', 0):.0%}  \n"
            md += f"- **Tags:** {', '.join(r.get('tags', []))}  \n\n"
            md += "```json\n"
            md += json.dumps(r.get("data", {}), indent=2, default=str)[:1000]
            md += "\n```\n\n---\n\n"

        md += "\n*Generated by [PhantomSignal OSINT Framework](https://github.com/getphantomsignal/phantomsignal)*"

        with open(path, "w", encoding="utf-8") as f:
            f.write(md)

    def _compress(self, path: Path) -> Path:
        zip_path = path.with_suffix(path.suffix + ".zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as zf:
            zf.write(path, path.name)
        return zip_path

    def _encrypt_file(self, path: Path, password: str) -> Path:
        """AES-256-GCM encryption via the cryptography library."""
        try:
            from cryptography.hazmat.primitives.ciphers.aead import AESGCM
            from cryptography.hazmat.primitives.kdf.scrypt import Scrypt
            import os

            salt = os.urandom(16)
            kdf = Scrypt(salt=salt, length=32, n=2**14, r=8, p=1)
            key = kdf.derive(password.encode())
            aesgcm = AESGCM(key)
            nonce = os.urandom(12)

            with open(path, "rb") as f:
                plaintext = f.read()

            ciphertext = aesgcm.encrypt(nonce, plaintext, None)
            enc_path = path.with_suffix(path.suffix + ".enc")

            with open(enc_path, "wb") as f:
                # Format: [4 bytes magic][16 bytes salt][12 bytes nonce][ciphertext]
                f.write(b"NPHM")
                f.write(salt)
                f.write(nonce)
                f.write(ciphertext)

            path.unlink(missing_ok=True)
            return enc_path
        except ImportError:
            logger.warning("cryptography library not installed — skipping encryption")
            return path

    def _compute_checksum(self, path: Path) -> str:
        sha256 = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                sha256.update(chunk)
        return sha256.hexdigest()

    def _human_size(self, size_bytes: int) -> str:
        for unit in ["B", "KB", "MB", "GB"]:
            if size_bytes < 1024:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024
        return f"{size_bytes:.1f} TB"

    @staticmethod
    def decrypt_file(enc_path: str, password: str, output_path: Optional[str] = None) -> str:
        """Decrypt an AES-256-GCM encrypted PhantomSignal export."""
        from cryptography.hazmat.primitives.ciphers.aead import AESGCM
        from cryptography.hazmat.primitives.kdf.scrypt import Scrypt

        with open(enc_path, "rb") as f:
            magic = f.read(4)
            if magic not in (b"NPHM", b"NOWL"):  # NOWL = legacy pre-1.4 files
                raise ValueError("Not a valid PhantomSignal encrypted file")
            salt = f.read(16)
            nonce = f.read(12)
            ciphertext = f.read()

        kdf = Scrypt(salt=salt, length=32, n=2**14, r=8, p=1)
        key = kdf.derive(password.encode())
        aesgcm = AESGCM(key)
        plaintext = aesgcm.decrypt(nonce, ciphertext, None)

        out = output_path or enc_path.replace(".enc", "")
        with open(out, "wb") as f:
            f.write(plaintext)
        return out
